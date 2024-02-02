# coding: utf-8
# Copyright (c) [2022] [federalsadler@sohu.com]
# [Clicker] is licensed under Mulan PSL v2.
# You can use this software according to the terms and conditions of the Mulan PSL v2.
# You may obtain a copy of Mulan PSL v2 at:
# http://license.coscl.org.cn/MulanPSL2
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
# EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
# MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
# See the Mulan PSL v2 for more details.
from __future__ import print_function

import datetime
import shutil

import openpyxl
from PyQt5 import QtCore
from PyQt5.QtCore import Qt, QUrl
from PyQt5.QtGui import QDesktopServices, QStandardItemModel, QStandardItem
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget,
                             QTableWidgetItem, QMessageBox, QHeaderView,
                             QDialog, QInputDialog, QMenu, QFileDialog)

from main_work import MainWork
from navigation import Na
from 设置窗口 import Setting
from 功能类 import exit_main_work
from 数据库操作 import *
from 窗体.about import Ui_Dialog
from 窗体.global_s import Ui_Global
from 窗体.info import Ui_Form
from 窗体.mainwindow import Ui_MainWindow
from 网页操作 import WebOption


# todo: 重写导航页功能类
# todo: 图片路径改用相对路径
# todo: 快捷键失效
# todo: 表格当前行插入指令
# todo: 导入指令可最近打开
# todo: 重新修改指令功能
# todo: 新增提示音功能
# todo: 新增倒计时窗口功能
# todo: 截图指令重新设计
# todo: 指令表格右键菜单
# todo: 指令执行使用多线程
# todo: 加载导航页窗口控件重新设计
# todo: 功能区功能重新设计
# todo: 终止循环功能
# todo: 执行指令改为多线程
# todo: OCR识别功能
# todo: 重写设置窗口

# activate clicker
# pyinstaller -F -w -i clicker.ico Clicker.py
# pyinstaller -D -w -i clicker.ico Clicker.py
# pyinstaller -D -i clicker.ico Clicker.py

# 添加指令的步骤：
# 1. 在导航页的页面中添加指令的控件
# 2. 在导航页的页面中添加指令的处理函数
# 3. 在导航页的treeWidget中添加指令的名称
# 4. 在功能类中添加运行功能


class Main_window(QMainWindow, Ui_MainWindow):
    """主窗口"""

    def __init__(self):
        super().__init__()
        # 初始化窗体
        self.setupUi(self)
        # 设置表格列宽自动变化，并使第5列列宽固定
        # self.format_table()
        # 窗口和信息
        self.version = 'v0.21'  # 软件版本
        self.navigation = Na(self)  # 实例化导航页窗口
        self.global_s = Global_s(self)  # 全局设置窗口
        self.main_work = MainWork(self, self.navigation)  # 窗体的功能
        self.setting = Setting(self)  # 实例化设置窗口Y
        self.about = About()  # 设置关于窗体
        self.info = Info()  # 运行提示窗口
        self.web_option = WebOption(self, self.navigation)  # 网页操作模块
        # 显示导不同的窗口
        self.pushButton.clicked.connect(lambda: self.show_windows('导航'))  # 显示导航窗口
        self.pushButton_3.clicked.connect(lambda: self.show_windows('全局'))  # 显示全局参数窗口
        self.actions_2.triggered.connect(lambda: self.show_windows('设置'))  # 打开设置
        self.actionabout.triggered.connect(lambda: self.show_windows('关于'))  # 打开关于窗体
        self.actionhelp.triggered.connect(lambda: self.show_windows('说明'))  # 打开使用说明
        # 主窗体表格功能
        self.toolButton_5.clicked.connect(self.get_data)  # 获取数据，主窗体刷新按钮
        self.navigation.pushButton_3.clicked.connect(self.get_data)  # 获取数据，子窗体取消按钮
        self.navigation.pushButton_2.clicked.connect(self.get_data)  # 获取数据，子窗体保存按钮
        self.pushButton_2.clicked.connect(self.delete_data)  # 删除数据，删除按钮
        self.toolButton_3.clicked.connect(lambda: self.go_up_down("up"))  # 交换数据，上移按钮
        self.toolButton_4.clicked.connect(lambda: self.go_up_down("down"))  # 交换数据，下移按钮
        self.actionb.triggered.connect(lambda: self.save_data_to_current('db'))  # 导出数据，导出按钮
        self.actiona.triggered.connect(lambda: self.save_data_to_current('excel'))
        self.toolButton_6.clicked.connect(self.clear_table)  # 清空指令
        self.actionf.triggered.connect(self.data_import)  # 导入数据
        self.pushButton_8.clicked.connect(self.modify_parameters)  # 修改参数按钮
        # 主窗体开始按钮
        self.pushButton_5.clicked.connect(self.start)
        self.pushButton_4.clicked.connect(lambda: self.start(only_current_instructions=True))
        # self.pushButton_6.clicked.connect(exit_main_work)  # 结束任务按钮
        self.toolButton_8.clicked.connect(self.exporting_operation_logs)  # 导出日志按钮
        # self.actionj.triggered.connect(lambda: self.check_update(1)) # 检查更新按钮（菜单栏）
        self.actiong.triggered.connect(self.hide_toolbar)  # 隐藏工具栏
        # 分支表名
        self.branch_name = []
        self.load_branch()
        # 创建和删除分支
        self.toolButton_2.clicked.connect(self.create_branch)
        self.toolButton.clicked.connect(self.delete_branch)
        self.comboBox.currentIndexChanged.connect(self.get_data)
        # 右键菜单
        self.tableWidget.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.tableWidget.customContextMenuRequested.connect(self.generateMenu)

    def format_table(self):
        """设置主窗口表格格式"""
        # 列的大小拉伸，可被调整
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # 列的大小为可交互式的，用户可以调整
        self.tableWidget.horizontalHeader().setSectionResizeMode(0, QHeaderView.Interactive)
        self.tableWidget.horizontalHeader().setSectionResizeMode(1, QHeaderView.Interactive)
        self.tableWidget.horizontalHeader().setSectionResizeMode(2, QHeaderView.Interactive)
        self.tableWidget.horizontalHeader().setSectionResizeMode(3, QHeaderView.Interactive)
        # 列的大小调整为固定，列宽不会改变
        self.tableWidget.horizontalHeader().setSectionResizeMode(6, QHeaderView.Fixed)
        self.tableWidget.horizontalHeader().setSectionResizeMode(7, QHeaderView.Fixed)
        # 设置列宽为50像素
        self.tableWidget.setColumnWidth(6, 60)
        self.tableWidget.setColumnWidth(7, 60)

    def generateMenu(self, pos):
        # 获取点击行号
        row_num = -1
        for i in self.tableWidget.selectionModel().selection().indexes():
            row_num = i.row()
        if row_num != -1:  # 未选中数据不弹出右键菜单
            menu = QMenu()  # 实例化菜单
            up_ins = menu.addAction("上移")
            down_ins = menu.addAction("下移")
            menu.addSeparator()
            copy_ins = menu.addAction("复制指令")
            modify_ins = menu.addAction("修改指令")
            del_ins = menu.addAction("删除指令")
            action = menu.exec_(self.tableWidget.mapToGlobal(pos))
        else:
            return
        # 各项操作
        if action == copy_ins:
            print("清除表格内容")
        elif action == del_ins:
            self.delete_data()
        elif action == up_ins:
            self.go_up_down('up')
        elif action == down_ins:
            self.go_up_down('down')
        elif action == modify_ins:
            self.modify_parameters()

    def show_windows(self, judge):
        """打开窗体"""
        resize = self.geometry()
        if judge == '设置':
            self.setting.show()
            # self.setting.load_setting_data()
            # self.setting.move(resize.x() + 90, resize.y())
        elif judge == '全局':
            self.global_s.setModal(True)
            self.global_s.exec_()
            # self.global_s.move(resize.x() + 90, resize.y())
        elif judge == '导航':
            self.navigation.show()
        elif judge == '关于':
            self.about.show()
            self.about.move(resize.x() + 90, resize.y())
        elif judge == '说明':
            QDesktopServices.openUrl(QUrl('https://gitee.com/automatic_clicker/automatic_clicker_2'))

    def get_data(self):
        """从数据库获取数据并存入表格"""
        print('刷新表格')
        try:
            self.tableWidget.clearContents()
            self.tableWidget.setRowCount(0)
            # 获取数据库数据
            cursor, con = sqlitedb()
            branch_name = self.comboBox.currentText()
            cursor.execute(
                'select 图像名称,指令类型,异常处理,备注,参数1,参数2,重复次数,ID from 命令 where 隶属分支=?',
                (branch_name,))
            list_order = cursor.fetchall()
            close_database(cursor, con)
            # 在表格中写入数据
            for i in range(len(list_order)):
                self.tableWidget.insertRow(i)
                for j in range(len(list_order[i])):
                    self.tableWidget.setItem(i, j, QTableWidgetItem(str(list_order[i][j])))
        except sqlite3.OperationalError:
            pass

    def delete_data(self):
        """删除选中的数据行"""
        # 获取选中值的行号和id
        try:
            row = self.tableWidget.currentRow()
            column = self.tableWidget.currentColumn()
            xx = self.tableWidget.item(row, 7).text()
            print(row, column, xx)
            # 删除数据库中指定id的数据
            con = sqlite3.connect('命令集.db')
            cursor = con.cursor()
            branch_name = self.comboBox.currentText()
            cursor.execute('delete from 命令 where ID=? and 隶属分支=?', (xx, branch_name,))
            con.commit()
            con.close()
            # 调用get_data()函数，刷新表格
            self.get_data()
        except AttributeError:
            pass

    def go_up_down(self, judge):
        """向上或向下移动选中的行"""

        def database_exchanges_two_rows(id_1: int, id_2: int) -> None:
            """交换数据库中的两行数据
            :param id_1: 要交换的第一行的id
            :param id_2: 要交换的第二行的id"""
            con = sqlite3.connect('命令集.db')
            cursor = con.cursor()
            # 交换两行的id
            cursor.execute('update 命令 set ID=? where ID=?', (999999, id_1,))
            cursor.execute('update 命令 set ID=? where ID=?', (id_1, id_2,))
            cursor.execute('update 命令 set ID=? where ID=?', (id_2, 999999,))
            con.commit()
            # 刷新数据库
            con.close()

        try:
            # 获取选中值的行号和id
            row = self.tableWidget.currentRow()
            id_ = int(self.tableWidget.item(row, 7).text())
            if judge == 'up':
                if row != 0:
                    # 查询上一行的id
                    id_row_up = int(self.tableWidget.item(row - 1, 7).text())
                    # 交换两行的id
                    database_exchanges_two_rows(id_, id_row_up)
                    self.get_data()
                    # 将焦点移动到交换后的行
                    self.tableWidget.setCurrentCell(row - 1, 0)
            elif judge == 'down':
                if row != self.tableWidget.rowCount() - 1:
                    # 查询下一行的id
                    id_row_down = int(self.tableWidget.item(row + 1, 7).text())
                    # 交换两行的id
                    database_exchanges_two_rows(id_, id_row_down)
                    self.get_data()
                    # 将焦点移动到交换后的行
                    self.tableWidget.setCurrentCell(row + 1, 0)
        except AttributeError:
            pass

    def save_data_to_current(self, judge):
        """保存配置文件到当前文件夹下
        :param judge: 保存的文件类型（db、excel）"""

        def get_instructions() -> list:
            """获取所有指令"""
            cursor, con_ = sqlitedb()
            cursor.execute('select * from 命令')
            list_instructions = cursor.fetchall()
            close_database(cursor, con_)
            return list_instructions

        def get_file_and_folder() -> tuple:
            """获取文件名和文件夹路径"""
            file_path = QFileDialog.getSaveFileName(
                self, "保存文件", '', "(*.db)" if judge == 'db' else "(*.xlsx)"
            )
            if file_path[0] != '':  # 获取文件名称
                return os.path.split(file_path[0])[1], os.path.split(file_path[0])[0]
            else:
                return None, None

        file_name, folder_path = get_file_and_folder()  # 获取文件名和文件夹路径
        if (file_name is not None) and (folder_path is not None):
            if judge == 'db':
                # 连接数据库
                con = sqlite3.connect('命令集.db')
                # 获取数据库文件路径
                db_file = con.execute('PRAGMA database_list').fetchall()[0][2]
                con.close()
                # 将数据库文件复制到指定文件夹下
                shutil.copy(db_file, folder_path + '/' + file_name)
            elif judge == 'excel':
                all_list_instructions = get_instructions()
                # 使用openpyxl模块创建Excel文件
                wb = openpyxl.Workbook()
                # 获取当前活动的sheet
                sheet = wb.active
                # 设置表头
                sheet['A1'] = 'ID'
                sheet['B1'] = '图像名称'
                sheet['C1'] = '指令类型'
                sheet['D1'] = '参数1'
                sheet['E1'] = '参数2'
                sheet['F1'] = '参数3'
                sheet['G1'] = '参数4'
                sheet['H1'] = '重复次数'
                sheet['I1'] = '异常处理'
                sheet['J1'] = '备注'
                sheet['K1'] = '隶属分支'
                # 写入数据
                for ins in range(len(all_list_instructions)):
                    for i in range(len(all_list_instructions[ins])):
                        sheet.cell(row=ins + 2, column=i + 1, value=all_list_instructions[ins][i])
                # 保存Excel文件
                wb.save(folder_path + '/' + file_name)
            QMessageBox.information(self, "提示", "指令数据保存成功！")

    @staticmethod
    def clear_database():
        """清空数据库"""
        cursor, con = sqlitedb()
        # 清空分支列表中所有的数据
        cursor.execute('delete from 命令 where ID<>-1')
        con.commit()
        close_database(cursor, con)

    def closeEvent(self, event):
        pass
        choice = QMessageBox.question(self, "提示", "确定退出并清空所有指令？")
        if choice == QMessageBox.Yes:
            # 退出终止后台进程并清空数据库
            event.accept()
            self.clear_database()
            self.web_option.close_browser()
            exit_main_work()
        else:
            event.ignore()

    def clear_table(self):
        """清空表格和数据库"""
        choice = QMessageBox.question(self, "提示", "确认清除所有指令吗？")
        if choice == QMessageBox.Yes:
            self.clear_database()
            self.get_data()
        else:
            pass

    def data_import(self):
        """导入数据功能"""
        # 打开选择文件对话框
        print('导入数据')
        target_path = QFileDialog.getOpenFileName(None, "请选择指令备份文件", '', "(*.db *.xlsx)")
        if target_path[0] == '':
            pass
        else:
            suffix = os.path.splitext(target_path[0])[1]
            # 如果为.db文件
            if suffix == '.db':
                # 获取当前文件夹路径
                # 将目标数据库中的数据导入到当前数据库中
                cursor, con = sqlitedb()
                # 获取目标数据库中的数据
                con_target = sqlite3.connect(target_path[0])
                cursor_target = con_target.cursor()
                cursor_target.execute('select * from 命令')
                list_instructions = cursor_target.fetchall()
                # 将数据导入到当前数据库中
                try:
                    for ins in list_instructions:
                        cursor.execute('insert into 命令 values (?,?,?,?,?,?,?,?,?,?,?)', ins)
                        con.commit()
                except sqlite3.IntegrityError:
                    QMessageBox.warning(self, "导入失败", "ID重复或格式错误！")
                    close_database(cursor, con)
                    return
                close_database(cursor, con)
                self.load_branch()
            # 如果为.xlsx文件
            elif suffix == '.xlsx':
                # 读取数据
                wb = openpyxl.load_workbook(target_path[0])
                sheet = wb.worksheets[0]
                max_row = sheet.max_row
                max_column = sheet.max_column
                # 连接数据库
                cursor, con = sqlitedb()
                try:
                    for row in range(2, max_row + 1):
                        # 获取第一列数据
                        instructions = []
                        for column in range(1, max_column + 1):
                            # 获取单元格数据
                            data = sheet.cell(row, column).value
                            instructions.append(data)
                        cursor.execute(
                            "INSERT INTO 命令(ID,图像名称,指令类型,参数1,参数2,参数3,参数4,"
                            "重复次数,异常处理,备注,隶属分支) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                            instructions[0:11])
                        con.commit()
                except Exception as e:
                    print(e)
                    QMessageBox.warning(self, "导入失败", "ID重复或格式错误！")
                    close_database(cursor, con)
                    return
                close_database(cursor, con)
                QMessageBox.information(self, "提示", "指令数据导入成功！")
                self.load_branch()

    def start(self, only_current_instructions=False):
        """主窗体开始按钮"""

        def info_show():
            """显示信息窗口"""
            self.info.show()
            resize = self.geometry()
            self.info.move(resize.x() + 45, resize.y() - 30)
            QApplication.processEvents()

        # 开始主任务
        if not only_current_instructions:
            info_show()
            self.main_work.start_work()
        elif only_current_instructions:
            if self.comboBox.currentText() == '主流程':
                QMessageBox.warning(self, "警告", "主分支无法执行该操作！")
            else:
                info_show()
                self.main_work.start_work(only_current_instructions)
        self.info.close()

    def main_show(self):
        """显示窗体，并根据设置检查更新"""
        self.show()
        # 连接数据库获取是否检查更新选项
        # con = sqlite3.connect('命令集.db')
        # cursor = con.cursor()
        # cursor.execute('select 值 from 设置 where 设置类型=?', ('启动检查更新',))
        # x = cursor.fetchall()[0][0]
        # cursor.close()
        # print('启动检查更新')
        # print(x)
        # if x == 1:
        #     self.check_update(0)
        # else:
        #     pass

    def hide_toolbar(self):
        """隐藏工具栏"""
        if self.actiong.isChecked():
            self.toolBar.show()
        elif not self.actiong.isChecked():
            self.toolBar.hide()

    def exporting_operation_logs(self):
        """导出操作日志"""
        # 打开保存文件对话框
        target_path = QFileDialog.getSaveFileName(self, "请选择保存路径", '', "(*.txt)")
        # 判断是否选择了文件
        if target_path[0] == '':
            pass
        else:
            # 获取操作日志
            logs = self.plainTextEdit.toPlainText()
            # 获取当前日期时间
            now = datetime.datetime.now()
            # 将操作日志写入文件
            with open(target_path[0], 'w') as f:
                f.write('日志导出时间：' + now.strftime('%Y-%m-%d %H:%M:%S') + '\n')
                f.write(logs)
            QMessageBox.information(self, "提示", "操作日志导出成功！")

    def modify_parameters(self):
        """修改参数"""
        try:
            # 获取当前行行号列号
            row = self.tableWidget.currentRow()
            id_ = self.tableWidget.item(row, 7).text()  # 指令ID
            ins_type = self.tableWidget.item(row, 1).text()  # 指令类型
            # 将导航页的tabWidget设置为对应的页
            self.show_windows('导航')
            self.navigation.switch_navigation_page(ins_type)
            # 修改数据中的参数
            self.navigation.modify_judgment = '修改'
            self.navigation.modify_id = id_
        except AttributeError:
            QMessageBox.information(self, "提示", "请先选择一行待修改的数据！")
            pass

    def create_branch(self):
        """创建分支表并重命名"""
        # 弹出输入对话框，提示输入分支名称
        text, ok = QInputDialog.getText(self, "创建分支", "请输入分支名称：")
        if ok:
            try:
                # 连接数据库
                cursor, con = sqlitedb()
                # 查找是否有同名分支
                cursor.execute('select 分支表名 from 全局参数 where 分支表名=?', (text,))
                x = cursor.fetchall()
                if len(x) > 0:
                    QMessageBox.information(self, "提示", "分支已存在！")
                    return
                else:
                    # 向全局参数表中添加分支表名
                    print('添加分支')
                    cursor.execute('insert into 全局参数(资源文件夹路径,分支表名) values(?,?)', (None, text))
                    con.commit()
                    # 弹出提示框，提示创建成功
                    QMessageBox.information(self, "提示", "分支创建成功！")
                # 关闭数据库连接
                close_database(cursor, con)
                # 加载分支
                self.load_branch()
            except sqlite3.OperationalError:
                QMessageBox.critical(self, "提示", "分支创建失败！")
                pass

    def delete_branch(self):
        """删除分支"""
        # 弹出输入对话框，提示输入分支名称
        print('删除分支')
        cursor, con = sqlitedb()
        text = self.comboBox.currentText()
        if text == '主流程':
            QMessageBox.information(self, "提示", "无法删除主分支！")
        else:
            # 将combox显示的名称切换为命令
            self.comboBox.setCurrentText('主流程')
            # 删除分支名称
            cursor.execute('delete from 全局参数 where 分支表名=?', (text,))
            # 关闭数据库连接
            con.commit()
            close_database(cursor, con)
            # 将分支名从分支列表中删除
            self.branch_name.remove(text)
            # 弹出提示框
            QMessageBox.information(self, "提示", "分支删除成功！")
            # 重新加载分支列表
            self.load_branch()

    def load_branch(self):
        """加载分支"""
        # 初始化功能
        print('加载分支')
        cursor, con = sqlitedb()
        # 获取所有分支名
        cursor.execute("select 分支表名 from 全局参数")
        self.branch_name = [x[0] for x in cursor.fetchall() if x[0] is not None]
        # 关闭数据库连接
        close_database(cursor, con)
        self.comboBox.clear()
        self.comboBox.addItems(self.branch_name)


class About(QWidget, Ui_Dialog):
    """关于窗体"""

    def __init__(self):
        super(About, self).__init__()
        self.setupUi(self)
        # 去除窗体最大化、最小化按钮
        self.setWindowFlags(Qt.WindowCloseButtonHint)
        self.setWindowModality(Qt.ApplicationModal)

        self.github.clicked.connect(self.show_github)
        self.gitee.clicked.connect(self.show_gitee)

    def show_github(self):
        # 弹出对话框显示“暂无信息”
        QMessageBox.information(self, '提醒', '暂无信息')

    @staticmethod
    def show_gitee():
        QDesktopServices.openUrl(QUrl('https://gitee.com/fasterthanlight/automatic_clicker'))


class Info(QDialog, Ui_Form):
    def __init__(self, parent=None):
        super(Info, self).__init__(parent)
        self.setupUi(self)
        self.setWindowModality(Qt.ApplicationModal)
        # 去除最大化最小化按钮
        self.setWindowFlags(Qt.WindowCloseButtonHint)
        self.setWindowModality(Qt.ApplicationModal)


class Global_s(QDialog, Ui_Global):
    """全局参数设置窗体"""

    def __init__(self, parent=None):
        super().__init__(parent)

        self.setupUi(self)
        # 去除帮助按钮
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.refresh_listview()  # 刷新listview
        self.pushButton.clicked.connect(self.select_file)  # 添加图像文件夹路径
        self.pushButton_2.clicked.connect(self.delete_listview)  # 删除listview中的项

    def select_file(self):
        """打开选择文件窗口,并将路径写入数据库"""
        fil_path = QFileDialog.getExistingDirectory(self, "选择存储目标图像的文件夹")
        if fil_path != '':
            global_write_to_database(os.path.normpath(fil_path))
        self.refresh_listview()

    def delete_listview(self):
        """删除listview中选中的那行数据"""

        # 获取选中的行的值
        def delete_data(value_):
            """删除数据库中的数据"""
            # 连接数据库
            cursor, conn = sqlitedb()
            cursor.execute("DELETE FROM 全局参数 WHERE 资源文件夹路径 = ?", (value_,))  # 删除数据
            cursor.execute("DELETE FROM 全局参数 WHERE 资源文件夹路径 is NULL and 分支表名 is Null")  # 删除无用数据条
            conn.commit()
            close_database(cursor, conn)

        try:
            indexes = self.listView.selectedIndexes()
            value = self.listView.model().itemFromIndex(indexes[0]).text()
            delete_data(value)  # 删除数据库中的数据
            self.refresh_listview()  # 刷新listview
        except Exception as e:
            print(e)

    def refresh_listview(self):
        """刷新listview"""

        def add_listview(list_, listview):
            """添加listview"""
            model = QStandardItemModel()
            listview.setModel(model)
            for item in list_:
                model.appendRow(QStandardItem(item))

        res_folder_path = extract_global_parameter('资源文件夹路径')  # 获取数据库中的数据
        add_listview(res_folder_path, self.listView)


if __name__ == "__main__":
    # 自适应高分辨率
    QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)
    app = QApplication([])
    main_window = Main_window()  # 创建主窗体
    main_window.main_show()  # 显示窗体，并根据设置检查更新
    sys.exit(app.exec_())

    # def is_admin():
    #     try:
    #         return ctypes.windll.shell32.IsUserAnAdmin()
    #     except:
    #         return False
    #
    # if is_admin():
    #     app = QApplication([])
    #     # 创建主窗体
    #     main_window_ = Main_window()
    #     # 显示窗体，并根据设置检查更新
    #     main_window_.main_show()
    #     # 显示添加对话框窗口
    #     sys.exit(app.exec_())
    # else:
    #     if sys.version_info[0] == 3:
    #         ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, __file__, None, 1)
    #     else:  # in python2.x
    #         ctypes.windll.shell32.ShellExecuteW(None, u"runas", unicode(sys.executable), unicode(__file__), None, 1)
